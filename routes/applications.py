import hashlib
import io
import os
from datetime import datetime
from html import escape as html_escape

from flask import Blueprint, jsonify, request, send_file, session

import ledger
from db import db_connect
from routes.auth import login_required

applications_bp = Blueprint('applications', __name__)

SUPPORTED_EXT = {
    '.png': 'image', '.jpg': 'image', '.jpeg': 'image',
    '.gif': 'image', '.bmp': 'image', '.webp': 'image',
    '.xlsx': 'excel',
    '.docx': 'word',
    '.pdf': 'pdf',
}

MAX_EXCEL_ROWS = 500


def _file_type(name):
    ext = os.path.splitext(name)[1].lower()
    return SUPPORTED_EXT.get(ext)


def _file_mtime_iso(filepath):
    try:
        return datetime.fromtimestamp(os.path.getmtime(filepath)).isoformat()
    except OSError:
        return None


def _file_sha256(filepath):
    try:
        h = hashlib.sha256()
        with open(filepath, 'rb') as f:
            for chunk in iter(lambda: f.read(1 << 20), b''):
                h.update(chunk)
        return h.hexdigest()
    except OSError:
        return None



@applications_bp.route('/api/applications/files')
@login_required
def list_files():
    folder = (request.args.get('folder') or '').strip()
    if not folder:
        return jsonify({'error': 'folder parameter required'}), 400
    if not os.path.isdir(folder):
        return jsonify({'error': 'not a valid directory'}), 400
    try:
        entries = os.listdir(folder)
    except PermissionError:
        return jsonify({'error': 'permission denied'}), 403

    disk_files = {}
    for name in sorted(entries):
        ftype = _file_type(name)
        if ftype and os.path.isfile(os.path.join(folder, name)):
            disk_files[name] = ftype

    with db_connect() as conn:
        rows = conn.execute(
            'SELECT id, filename, status, file_mtime FROM app_file_status '
            'WHERE folder=? ORDER BY created_at ASC', (folder,)
        ).fetchall()

    dismissed = set()
    flagged = {}
    rejected = []
    approved = {}
    for r in rows:
        if r['status'] == 'dismissed':
            dismissed.add(r['filename'])
        elif r['status'] == 'flagged':
            flagged[r['filename']] = {'id': r['id'], 'mtime': r['file_mtime']}
        elif r['status'] == 'rejected':
            rejected.append({'id': r['id'], 'filename': r['filename'], 'mtime': r['file_mtime']})
        elif r['status'] == 'approved':
            # rows are ordered created_at ASC, so the latest approval's mtime wins
            approved[r['filename']] = r['file_mtime']

    stale_flag_ids = []
    for name, info in list(flagged.items()):
        if name in disk_files:
            current_mtime = _file_mtime_iso(os.path.join(folder, name))
            if current_mtime and current_mtime != info['mtime']:
                stale_flag_ids.append(info['id'])
                del flagged[name]
    if stale_flag_ids:
        with db_connect() as conn:
            for sid in stale_flag_ids:
                conn.execute('DELETE FROM app_file_status WHERE id=?', (sid,))

    with ledger.ledger_connect() as lconn:
        approved_hashes = ledger.latest_hashes_for_folder(lconn, folder)

    files = []
    handled_by_reject = set()

    for rej in rejected:
        name = rej['filename']
        ftype = _file_type(name) or 'pdf'
        if name in disk_files:
            current_mtime = _file_mtime_iso(os.path.join(folder, name))
            if current_mtime == rej['mtime']:
                handled_by_reject.add(name)
            else:
                files.append({'name': name, 'type': ftype, 'status': 'rejected', 'ghost': True})
        else:
            files.append({'name': name, 'type': ftype, 'status': 'rejected', 'ghost': True})

    for name, ftype in disk_files.items():
        if name in dismissed:
            continue
        if name in handled_by_reject:
            files.append({'name': name, 'type': ftype, 'status': 'rejected'})
            continue
        entry = {'name': name, 'type': ftype}
        if name in approved:
            if name in approved_hashes:
                # content hash from the ledger is the source of truth — immune
                # to mtime spoofing, identical for every supported file type
                unchanged = _file_sha256(os.path.join(folder, name)) == approved_hashes[name]
            else:
                # legacy approval predating the ledger (unmigrated): mtime fallback
                unchanged = _file_mtime_iso(os.path.join(folder, name)) == approved[name]
            if unchanged:
                entry['status'] = 'approved'
            elif name in flagged:
                entry['status'] = 'flagged'
            else:
                # approved earlier but changed since — needs a fresh approval
                entry['status'] = 'modified'
        elif name in flagged:
            entry['status'] = 'flagged'
        files.append(entry)

    files.sort(key=lambda f: (f.get('ghost', False), f['name'].lower()))
    return jsonify({'files': files, 'folder': folder})


@applications_bp.route('/api/applications/file/status', methods=['POST'])
@login_required
def set_file_status():
    data = request.json or {}
    folder = (data.get('folder') or '').strip()
    filename = (data.get('filename') or '').strip()
    status = (data.get('status') or '').strip()
    if not folder or not filename or status not in ('dismissed', 'flagged', 'rejected', 'approved'):
        return jsonify({'error': 'folder, filename, and valid status required'}), 400

    filepath = os.path.join(folder, filename)
    mtime = _file_mtime_iso(filepath)

    approved_path = None
    if status == 'approved':
        if not os.path.isfile(filepath):
            return jsonify({'error': 'file not found'}), 404
        grandparent = os.path.basename(os.path.dirname(folder))
        parent = os.path.basename(folder)
        if not grandparent:
            grandparent = '_root'
        # '/' regardless of platform: approved_path is a ledger key, not a filesystem path
        approved_path = '/'.join([grandparent, parent, filename])
        try:
            with open(filepath, 'rb') as f:
                content = f.read()
        except OSError:
            return jsonify({'error': 'cannot read file'}), 400
        with ledger.ledger_connect() as lconn:
            ledger.append_entry(
                lconn, folder, filename, approved_path, content,
                file_mtime=mtime, approved_by=session.get('username'),
            )

    with db_connect() as conn:
        if status in ('dismissed', 'flagged'):
            conn.execute(
                'DELETE FROM app_file_status WHERE folder=? AND filename=? AND status IN (?, ?)',
                (folder, filename, 'dismissed', 'flagged')
            )
        conn.execute(
            'INSERT INTO app_file_status (folder, filename, status, file_mtime, created_at, approved_path) '
            'VALUES (?, ?, ?, ?, ?, ?)',
            (folder, filename, status, mtime, datetime.now().isoformat(), approved_path)
        )
    return jsonify({'ok': True})


@applications_bp.route('/api/applications/approved-tree')
@login_required
def approved_tree():
    with ledger.ledger_connect() as lconn:
        latest = ledger.latest_entries_by_path(lconn)

    grouped = {}
    for path, row in latest.items():
        parts = path.split('/')
        if len(parts) != 3:
            continue
        gp_name, p_name, fname = parts
        ftype = _file_type(fname)
        if not ftype:
            continue
        grouped.setdefault(gp_name, {}).setdefault(p_name, []).append(
            {'name': fname, 'type': ftype, 'approved_at': row['created_at']}
        )

    tree = []
    for gp_name in sorted(grouped):
        gp_node = {'name': gp_name, 'children': [], 'last_updated': None}
        for p_name in sorted(grouped[gp_name]):
            children = sorted(grouped[gp_name][p_name], key=lambda f: f['name'])
            p_node = {'name': p_name, 'children': children, 'last_updated': None}
            for child in children:
                a = child['approved_at']
                if a and (not p_node['last_updated'] or a > p_node['last_updated']):
                    p_node['last_updated'] = a
            gp_node['children'].append(p_node)
            if p_node['last_updated'] and (not gp_node['last_updated'] or p_node['last_updated'] > gp_node['last_updated']):
                gp_node['last_updated'] = p_node['last_updated']
        tree.append(gp_node)
    return jsonify({'tree': tree})


@applications_bp.route('/api/applications/ledger/verify')
@login_required
def ledger_verify():
    with ledger.ledger_connect() as lconn:
        result = ledger.verify_chain(lconn)
    return jsonify(result)


def _safe_path(folder, name):
    joined = os.path.join(folder, name)
    real = os.path.realpath(joined)
    if not real.startswith(os.path.realpath(folder)):
        return None
    return real


@applications_bp.route('/api/applications/file/preview')
@login_required
def file_preview():
    folder = (request.args.get('folder') or '').strip()
    name = (request.args.get('name') or '').strip()
    if not folder or not name:
        return jsonify({'error': 'folder and name required'}), 400

    ftype = _file_type(name)
    if not ftype:
        return jsonify({'error': 'unsupported file type'}), 400

    if request.args.get('approved') == '1':
        # approved previews are served from the verified blob store, never disk
        approved_path = folder + '/' + name
        with ledger.ledger_connect() as lconn:
            entry = ledger.latest_entry_for_path(lconn, approved_path)
            content = ledger.fetch_blob(lconn, entry['content_sha256']) if entry else None
        if entry is None or content is None:
            return jsonify({'error': 'no approved record for this file'}), 404
        if hashlib.sha256(content).hexdigest() != entry['content_sha256']:
            return jsonify({'error': 'integrity check failed - stored content does not match the ledger'}), 409
        return _dispatch_preview(ftype, io.BytesIO(content), name)

    filepath = _safe_path(folder, name)
    if filepath is None:
        return jsonify({'error': 'forbidden'}), 403
    if not os.path.isfile(filepath):
        return jsonify({'error': 'file not found'}), 404
    return _dispatch_preview(ftype, filepath, name)


def _dispatch_preview(ftype, source, name):
    """source is a filesystem path (live files) or BytesIO (approved blobs)."""
    if ftype == 'image':
        return _preview_image(source, name)
    if ftype == 'pdf':
        return _preview_pdf(source)
    if ftype == 'excel':
        page = request.args.get('page', '0')
        try:
            page = max(0, int(page))
        except ValueError:
            page = 0
        return _preview_excel(source, page)
    if ftype == 'word':
        return _preview_word(source)
    return jsonify({'error': 'unsupported'}), 400


def _preview_image(source, name):
    import mimetypes
    mime = mimetypes.guess_type(name)[0] or 'application/octet-stream'
    return send_file(source, mimetype=mime)


def _preview_pdf(source):
    return send_file(source, mimetype='application/pdf')


def _preview_excel(source, page):
    import openpyxl
    try:
        wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    except Exception:
        return jsonify({'error': 'cannot read Excel file'}), 400
    try:
        sheets = wb.sheetnames
        if not sheets:
            return jsonify({'error': 'no sheets found'}), 400
        page = min(page, len(sheets) - 1)
        ws = wb[sheets[page]]

        # openpyxl's reported dimension is often an inflated "used range"
        # (trailing blank rows/columns left by past edits or formatting), and
        # every row is padded out to max_column. Trim to the real data bounding
        # box: drop rows after the last one holding a value, and columns after
        # the rightmost non-empty cell across all rows.
        rows = []
        last_data_row = -1
        ncols = 0
        truncated = False
        for r in ws.iter_rows(values_only=True):
            if len(rows) >= MAX_EXCEL_ROWS:
                if any(cell is not None for cell in r):
                    truncated = True
                    break
                continue
            row = []
            row_width = 0
            for idx, cell in enumerate(r):
                if cell is None:
                    row.append('')
                else:
                    row_width = idx + 1
                    if isinstance(cell, datetime):
                        row.append(cell.isoformat())
                    else:
                        row.append(str(cell))
            rows.append(row)
            if row_width:
                last_data_row = len(rows) - 1
                if row_width > ncols:
                    ncols = row_width

        if last_data_row < 0:
            headers, data_rows = [], []
        else:
            trimmed = []
            for row in rows[:last_data_row + 1]:
                row = row[:ncols]
                if len(row) < ncols:
                    row.extend([''] * (ncols - len(row)))
                trimmed.append(row)
            headers = trimmed[0]
            data_rows = trimmed[1:]

        return jsonify({
            'type': 'excel',
            'page': page,
            'total_pages': len(sheets),
            'page_label': sheets[page],
            'headers': headers,
            'rows': data_rows,
            'truncated': truncated,
        })
    finally:
        wb.close()


def _preview_word(source):
    import docx
    from docx.oxml.ns import qn
    try:
        doc = docx.Document(source)
    except Exception:
        return jsonify({'error': 'cannot read Word file'}), 400

    parts = []
    for child in doc.element.body:
        if child.tag == qn('w:p'):
            para = docx.text.paragraph.Paragraph(child, doc)
            parts.append(_paragraph_to_html(para))
        elif child.tag == qn('w:tbl'):
            table = docx.table.Table(child, doc)
            parts.append(_table_to_html(table))

    return jsonify({
        'type': 'word',
        'page': 0,
        'total_pages': 1,
        'html': '\n'.join(parts),
    })


def _paragraph_to_html(para):
    style_name = (para.style.name or '').lower()
    tag = 'p'
    for level in range(1, 7):
        if style_name == f'heading {level}':
            tag = f'h{level}'
            break

    runs_html = []
    for run in para.runs:
        text = html_escape(run.text)
        if run.bold:
            text = f'<strong>{text}</strong>'
        if run.italic:
            text = f'<em>{text}</em>'
        if run.underline:
            text = f'<u>{text}</u>'
        runs_html.append(text)

    inner = ''.join(runs_html)
    if not inner.strip():
        return '<p>&nbsp;</p>'
    return f'<{tag}>{inner}</{tag}>'


def _table_to_html(table):
    rows_html = []
    for row in table.rows:
        cells = ''.join(
            f'<td>{html_escape(cell.text)}</td>' for cell in row.cells
        )
        rows_html.append(f'<tr>{cells}</tr>')
    return f'<table class="app-docx-table">{"".join(rows_html)}</table>'
